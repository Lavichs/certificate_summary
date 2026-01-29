import json
import os
import uuid
import openpyxl
from datetime import datetime
from typing import Annotated
from pathlib import Path
from redis.asyncio import Redis

from fastapi import APIRouter, Depends, UploadFile, Body
from fastapi.encoders import jsonable_encoder

from config import settings
from src.schemas.certificate import SCertificateAdd
from src.services.certificate import CertificateService
from src.api.depends import certificate_service, get_session_data
from src.utils.redis_client import get_redis

router = APIRouter(
    prefix="/certificates",
    tags=["Certificates"],
)


@router.get("")
async def getSummary(
    certificate_service: Annotated[CertificateService, Depends(certificate_service)],
    redis: Redis = Depends(get_redis)
):
    cache_data = await redis.get("all_data")
    if cache_data is None:
        data = await certificate_service.getAll()
        await redis.set("all_data", json.dumps(jsonable_encoder(data)), ex=settings.CACHE_LIFETIME)
        return data
    return json.loads(cache_data)


@router.put("/{id}")
async def update(
    certificate_service: Annotated[CertificateService, Depends(certificate_service)],
    id: str,
    data=Body(),
    redis: Redis = Depends(get_redis)
):
    result = await certificate_service.update(id, data)
    data = await certificate_service.getAll()
    await redis.set("all_data", json.dumps(jsonable_encoder(data)), ex=settings.CACHE_LIFETIME)
    return result


@router.delete("/{id}")
async def delete_certificate(
    certificate_service: Annotated[CertificateService, Depends(certificate_service)],
    id: str,
    redis: Redis = Depends(get_redis)
):
    await certificate_service.delete(id)
    data = await certificate_service.getAll()
    await redis.set("all_data", json.dumps(jsonable_encoder(data)), ex=settings.CACHE_LIFETIME)
    return {"status": "success"}


@router.post("/loadxlsx")
async def updateDatabaseByXlsx(
    uploaded_file: UploadFile,
    certificate_service: Annotated[CertificateService, Depends(certificate_service)],
    redis: Redis = Depends(get_redis),
    user_session_data: dict = Depends(get_session_data),
):
    upload_folder_path = Path(__file__).resolve().parent.parent.parent.parent
    path_to_file = Path(
        upload_folder_path,
        "upload_files",
        f"{uuid.uuid4()}.{uploaded_file.filename.split('.')[-1]}",
    )
    with open(path_to_file, "wb") as f:
        f.write(uploaded_file.file.read())

    await certificate_service.deleteAll()

    wb_obj = openpyxl.load_workbook(path_to_file)
    sheet_obj = wb_obj.active
    for i in range(2, sheet_obj.max_row):
        cert_center = sheet_obj.cell(row=i, column=1).value
        fio = sheet_obj.cell(row=i, column=2).value
        post = sheet_obj.cell(row=i, column=3).value
        organization = sheet_obj.cell(row=i, column=4).value
        expire_date = sheet_obj.cell(row=i, column=5).value
        if isinstance(expire_date, str):
            expire_date = datetime.strptime(expire_date, "%d.%m.%Y").date()
        status = sheet_obj.cell(row=i, column=6).value
        comment = sheet_obj.cell(row=i, column=7).value

        scsc = SCertificateAdd(
            cert_center=cert_center,
            fio=fio,
            post=post,
            organization=organization,
            expire_date=expire_date,
            status=status,
            comment=str(comment),
        )
        await certificate_service.create(scsc)
    os.remove(path_to_file)

    data = await certificate_service.getAll()
    await redis.set("all_data", json.dumps(jsonable_encoder(data)), ex=settings.CACHE_LIFETIME)

    return {"status": "upload"}
